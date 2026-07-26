"""
Heuristic parser: a FOIA-reply attachment (or email body) -> normalized
permit dicts in the exact shape app/ingest.py's upsert_permit expects.

Why heuristic: unlike the API connectors, there is no metadata endpoint
and no stable schema here. Every agency exports its own idiosyncratic
column names ("Permit No" vs "PERMIT_NUMBER" vs "Application #"), so this
module maps *observed* column headers onto our canonical fields by
fuzzy, case-insensitive keyword matching (FIELD_KEYWORDS below), using a
"longest matching keyword wins" rule so a specific header ("estimated
cost") beats a generic one ("cost") when both could match.

Design commitments:
  * Nothing is fabricated. A column only maps if its header matches a
    known keyword; unmatched columns are NOT discarded -- the entire
    original row is preserved verbatim in ``raw_data`` so nothing is lost
    and a better mapping can be re-derived later.
  * Every record is flagged ``needs_review=True``.
  * PDFs are best-effort. If pdfplumber can't find a clean table, we fail
    gracefully (status="unparseable") and flag it for a human rather than
    inventing garbage rows.
  * permit_number is required by the upsert. When a row genuinely has no
    identifiable permit number, a deterministic synthetic id is minted
    from a hash of the row so re-parsing the same data is idempotent.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional

from app.connectors.base import NORMALIZED_PERMIT_FIELDS

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Field-mapping heuristics
# ---------------------------------------------------------------------------
# canonical field -> list of keyword variants (matched case-insensitively
# against a normalized header). Keywords <= 3 chars must appear as a
# standalone token (avoids "lat" matching "plate"); longer keywords match
# as substrings. "Longest matching keyword wins" resolves ambiguity across
# fields (e.g. "estimated cost" -> estimated_cost, not valuation's "value").
FIELD_KEYWORDS: dict[str, list[str]] = {
    "permit_number": [
        "permit number", "permit no", "permit num", "permit id", "permit #",
        "permitnumber", "permit_number", "permit", "record id", "record number",
        "application number", "application no", "application #", "app no", "app #",
        "case number", "case no", "bp number", "bp no", "folio", "reference number",
    ],
    "permit_type": [
        "permit type", "type of permit", "permit class", "record type",
        "application type", "permittype", "permit_type",
    ],
    "status": ["permit status", "current status", "status", "disposition"],
    "application_date": [
        "application date", "applied date", "date applied", "app date",
        "filed date", "date filed", "submitted date", "date submitted",
        "submittal date", "date received", "received date", "intake date",
    ],
    "issue_date": [
        "issue date", "issued date", "date issued", "issuance date",
        "issuedate", "issued", "issue",
    ],
    "completion_date": [
        "completion date", "completed date", "date completed", "finaled date",
        "final date", "date finaled", "co date", "certificate of occupancy",
        "complete",
    ],
    "expiration_date": ["expiration date", "expires date", "date expires", "expiration", "expires", "expire"],
    "contractor": [
        "general contractor", "contractor company", "contractor name",
        "contractor", "gc name", "gc",
    ],
    "builder": ["builder", "applicant name", "applicant", "owner name", "owner"],
    "architect": ["architect name", "architect"],
    "engineer": ["engineer name", "engineer"],
    "property_address": [
        "property address", "site address", "project address", "job address",
        "full address", "street address", "location address", "address",
        "location", "site", "premise",
    ],
    "parcel_number": [
        "parcel number", "parcel no", "parcel id", "assessor parcel",
        "apn", "parcel", "pin", "tax map", "tax id", "gpin", "folio number",
    ],
    "estimated_cost": [
        "estimated cost", "est cost", "estimated construction cost",
        "construction cost", "project cost", "job cost", "job value",
        "cost of construction", "cost",
    ],
    "valuation": [
        "total valuation", "declared valuation", "declared value",
        "estimated value", "job valuation", "valuation", "value",
    ],
    "description": [
        "work description", "project description", "description of work",
        "scope of work", "description", "scope", "work desc", "work",
    ],
    "work_category": [
        "work category", "work type", "work class", "classification",
        "use group", "use type", "category", "use",
    ],
    "square_footage": [
        "square footage", "square feet", "sq ft", "sq. ft.", "sqft",
        "gross square", "floor area", "total area", "area",
    ],
    "units": ["dwelling units", "number of units", "num units", "no of units", "units"],
    "latitude": ["latitude", "lat", "y coord", "ycoord"],
    "longitude": ["longitude", "long", "lon", "lng", "x coord", "xcoord"],
}

# Fields whose mapped value should be coerced through a datetime parser.
_DATE_FIELDS = {"application_date", "issue_date", "completion_date", "expiration_date"}
_FLOAT_FIELDS = {"estimated_cost", "valuation", "square_footage", "latitude", "longitude"}
_INT_FIELDS = {"units"}


@dataclass
class ParseResult:
    status: str  # "parsed" | "no_records" | "unsupported" | "unparseable"
    records: list[dict] = field(default_factory=list)
    note: str = ""
    detected_columns: list[str] = field(default_factory=list)
    field_mapping: dict[str, str] = field(default_factory=dict)  # canonical field -> source column

    @property
    def ok(self) -> bool:
        return self.status == "parsed" and bool(self.records)


# ---------------------------------------------------------------------------
# Header -> canonical field mapping
# ---------------------------------------------------------------------------


def _normalize_header(header: str) -> str:
    h = (header or "").strip().lower()
    h = re.sub(r"[_\-/]+", " ", h)
    h = re.sub(r"[^a-z0-9# ]+", " ", h)
    h = re.sub(r"\s+", " ", h).strip()
    return h


def _keyword_score(norm_header: str, keyword: str) -> int:
    """Return the match strength (keyword length) or 0 if no match."""
    if not keyword:
        return 0
    if len(keyword) <= 3:
        # short keyword must be a standalone token
        return len(keyword) if keyword in norm_header.split() else 0
    return len(keyword) if keyword in norm_header else 0


def build_field_mapping(headers: list[str]) -> dict[str, str]:
    """
    Decide, for each canonical field, which source column best fills it.

    Every (canonical field, column) pair is scored by its strongest
    keyword match; each canonical field claims its highest-scoring column,
    and no column is assigned to two different fields (the stronger match
    wins the column). Returns {canonical_field: source_column_header}.
    """
    norm = {h: _normalize_header(h) for h in headers}

    # candidate[(field, header)] = score
    candidates: list[tuple[int, str, str]] = []
    for field_name, keywords in FIELD_KEYWORDS.items():
        for header in headers:
            best = max((_keyword_score(norm[header], kw) for kw in keywords), default=0)
            if best > 0:
                candidates.append((best, field_name, header))

    # Highest score first; assign greedily so both a field and a column are
    # used at most once.
    candidates.sort(key=lambda c: c[0], reverse=True)
    mapping: dict[str, str] = {}
    used_columns: set[str] = set()
    for _score, field_name, header in candidates:
        if field_name in mapping or header in used_columns:
            continue
        mapping[field_name] = header
        used_columns.add(header)
    return mapping


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------


def _to_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[,$\s]", "", str(value))
    if text in ("", "-", "n/a", "na", "none"):
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _to_int(value: Any) -> Optional[int]:
    f = _to_float(value)
    return int(f) if f is not None else None


_DATE_FORMATS = (
    "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y/%m/%d",
    "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y", "%d-%b-%Y", "%d-%b-%y",
    "%b %d, %Y", "%B %d, %Y", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
)


def _to_datetime(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        logger.debug("FOIA parser: could not parse date value %r", value)
        return None


def _coerce(field_name: str, value: Any) -> Any:
    if field_name in _DATE_FIELDS:
        return _to_datetime(value)
    if field_name in _FLOAT_FIELDS:
        return _to_float(value)
    if field_name in _INT_FIELDS:
        return _to_int(value)
    if value is None:
        return None
    text = str(value).strip()
    return text or None


# ---------------------------------------------------------------------------
# Row -> normalized permit dict
# ---------------------------------------------------------------------------


def _synthetic_permit_number(raw_row: dict, source_label: str) -> str:
    """Deterministic id from the row content, so re-parsing is idempotent."""
    blob = "|".join(f"{k}={v}" for k, v in sorted(raw_row.items()))
    digest = hashlib.sha1(f"{source_label}|{blob}".encode("utf-8", "replace")).hexdigest()[:16]
    return f"FOIA-{digest}"


def map_row(raw_row: dict, mapping: dict[str, str], source_label: str) -> dict:
    """Map one raw row into a normalized permit dict (raw_data preserved)."""
    normalized = {k: None for k in NORMALIZED_PERMIT_FIELDS}
    for field_name, source_col in mapping.items():
        if field_name in normalized:
            normalized[field_name] = _coerce(field_name, raw_row.get(source_col))

    # A cost/value column commonly fills both estimated_cost and valuation
    # in the API connectors; mirror that when only one side was detected.
    if normalized.get("valuation") is None and normalized.get("estimated_cost") is not None:
        normalized["valuation"] = normalized["estimated_cost"]
    if normalized.get("estimated_cost") is None and normalized.get("valuation") is not None:
        normalized["estimated_cost"] = normalized["valuation"]

    permit_number = normalized.get("permit_number")
    if not permit_number or not str(permit_number).strip():
        normalized["permit_number"] = _synthetic_permit_number(raw_row, source_label)
    else:
        normalized["permit_number"] = str(permit_number).strip()

    normalized["source"] = source_label
    normalized["needs_review"] = True
    # Preserve EVERY original column (mapped and unmapped) verbatim.
    normalized["raw_data"] = dict(raw_row)
    return normalized


def _records_from_rows(rows: list[dict], source_label: str) -> ParseResult:
    """Shared tail: given a list of dict rows, map them all."""
    rows = [r for r in rows if any(v not in (None, "") for v in r.values())]
    if not rows:
        return ParseResult(status="no_records", note="No data rows found.")
    headers = list(rows[0].keys())
    mapping = build_field_mapping(headers)
    if "permit_number" not in mapping and not mapping:
        # Nothing recognizable at all -- still ingest (synthetic ids) but flag it.
        note = "No recognizable permit columns; ingested with synthetic ids for review."
    else:
        note = ""
    records = [map_row(r, mapping, source_label) for r in rows]
    return ParseResult(
        status="parsed",
        records=records,
        note=note,
        detected_columns=headers,
        field_mapping=mapping,
    )


# ---------------------------------------------------------------------------
# Format-specific extractors
# ---------------------------------------------------------------------------


def parse_csv(data: bytes, source_label: str) -> ParseResult:
    text = _decode_text(data)
    if not text.strip():
        return ParseResult(status="no_records", note="Empty CSV.")
    # Sniff the delimiter (agencies export comma, tab, or semicolon).
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        return ParseResult(status="no_records", note="CSV had no header row.")
    rows = [dict(r) for r in reader]
    return _records_from_rows(rows, source_label)


def parse_xlsx(data: bytes, source_label: str) -> ParseResult:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        return ParseResult(status="unsupported", note="openpyxl not installed.")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        return ParseResult(status="unparseable", note=f"Could not open workbook: {exc}")

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        header_row = None
        for row in rows_iter:
            if row and any(c not in (None, "") for c in row):
                header_row = row
                break
        if header_row is None:
            continue
        headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(header_row)]
        data_rows: list[dict] = []
        for row in rows_iter:
            if row is None or all(c in (None, "") for c in row):
                continue
            data_rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        if data_rows:
            wb.close()
            return _records_from_rows(data_rows, source_label)
    wb.close()
    return ParseResult(status="no_records", note="No sheet with data rows found.")


def parse_pdf(data: bytes, source_label: str) -> ParseResult:
    try:
        import pdfplumber
    except ImportError:  # pragma: no cover
        return ParseResult(status="unsupported", note="pdfplumber not installed.")
    try:
        all_rows: list[dict] = []
        headers: Optional[list[str]] = None
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    if not table or len(table) < 2:
                        continue
                    raw_header = table[0]
                    if not any(c and str(c).strip() for c in raw_header):
                        continue
                    page_headers = [
                        (str(c).strip() if c and str(c).strip() else f"col_{i}")
                        for i, c in enumerate(raw_header)
                    ]
                    if headers is None:
                        headers = page_headers
                    for body_row in table[1:]:
                        if not any(c and str(c).strip() for c in body_row):
                            continue
                        all_rows.append(
                            {headers[i]: (body_row[i] if i < len(body_row) else None) for i in range(len(headers))}
                        )
        if not all_rows:
            return ParseResult(
                status="unparseable",
                note="PDF had no cleanly extractable table structure; flagged for manual review.",
            )
        return _records_from_rows(all_rows, source_label)
    except Exception as exc:
        return ParseResult(status="unparseable", note=f"PDF extraction failed: {exc}; flagged for manual review.")


def parse_email_body(text: str, source_label: str) -> ParseResult:
    """
    Best-effort: some agencies paste the data as a small table right in the
    reply body rather than attaching a file. Only treat the body as data if
    it clearly looks tabular (a delimiter-consistent header + rows);
    otherwise return no_records so a plain prose reply isn't mangled.
    """
    if not text or not text.strip():
        return ParseResult(status="no_records", note="Empty body.")
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return ParseResult(status="no_records", note="Body not tabular.")
    for delim in ("\t", "|", ","):
        header_count = lines[0].count(delim)
        if header_count >= 1 and all(ln.count(delim) >= header_count for ln in lines[1:4] if ln is not lines[0]):
            block = "\n".join(lines)
            result = parse_csv(block.encode("utf-8"), source_label)
            if result.ok:
                result.note = (result.note + " (parsed from inline email body)").strip()
                return result
    return ParseResult(status="no_records", note="Email body is prose, not a data table.")


# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------


def parse_attachment(filename: str, data: bytes, source_label: str) -> ParseResult:
    """Dispatch to the right extractor based on filename/content."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt") or name.endswith(".tsv"):
        return parse_csv(data, source_label)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(data, source_label)
    if name.endswith(".xls"):
        # Legacy binary .xls isn't supported by openpyxl; flag for a human
        # rather than silently dropping it.
        return ParseResult(
            status="unsupported",
            note="Legacy .xls format not supported (openpyxl reads .xlsx only); flagged for manual review.",
        )
    if name.endswith(".pdf"):
        return parse_pdf(data, source_label)
    # Unknown extension: sniff the bytes.
    if data[:4] == b"PK\x03\x04":  # zip container -> xlsx
        return parse_xlsx(data, source_label)
    if data[:5] == b"%PDF-":
        return parse_pdf(data, source_label)
    # Fall back to trying CSV (many agencies attach a .dat/.text CSV).
    result = parse_csv(data, source_label)
    if result.ok:
        return result
    return ParseResult(status="unsupported", note=f"Unrecognized attachment type: {filename!r}")


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")
